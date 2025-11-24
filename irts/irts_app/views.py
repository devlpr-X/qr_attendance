from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db.models import Count, Q
from datetime import datetime, date, timedelta
import csv
import io
import math
from math import radians, sin, cos, sqrt, asin

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import io

from .models import (
    ClassSession, Student, Enrollment, Attendance, 
    WeeklySchedule, Course, TeacherProfile, Location, AttendanceReport
)


# ============================================
# UTILITY FUNCTIONS
# ============================================

def haversine_distance(lat1, lon1, lat2, lon2):
    """
    Хоёр GPS координатын хоорондох зайг метрээр тооцоолох
    """
    R = 6371000.0  # Дэлхийн радиус метрээр
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = sin(dlat/2)**2 + cos(radians(lat1))*cos(radians(lat2))*sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    return R * c


# ============================================
# TEACHER DASHBOARD VIEWS
# ============================================
# irts_app/views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import ClassSession, AttendanceSession

@login_required
def teacher_dashboard(request):
    # Хэрэглэгч TeacherProfile-той эсэхийг шалгах
    if not hasattr(request.user, 'teacherprofile'):
        # Хэрэв байхгүй бол профайл үүсгэх хуудас руу шилжүүлэх эсвэл мессеж харуулах
        return render(request, 'teacher/no_profile.html')

    teacher = request.user.teacherprofile
    today = timezone.localdate()
    weekday = today.weekday()  # 0=Даваа, 6=Ням

    # Тухайн багшийн өнөөдрийн хичээлүүд
    today_sessions = ClassSession.objects.filter(
        teacher=teacher,
        weekday=weekday
    )

    # # Өнөөдөр эхэлсэн ирц бүртгэлийн сессууд
    # active_sessions = AttendanceSession.objects.filter(
    #     class_session__in=today_sessions,
    #     date=today,
    #     is_active=True
    # )

    context = {
        'today_sessions': today_sessions,
        # 'active_sessions': active_sessions,
        'today': today,
    }
    return render(request, 'teacher/dashboard.html', context)

@login_required
def select_current_class(request):
    """Одоогийн цагт орж байгаа хичээлээ сонгох"""
    try:
        teacher = request.user.teacherprofile
    except:
        return redirect('teacher_dashboard')
    
    today = date.today()
    today_weekday = today.weekday()
    
    # Өнөөдрийн хуваарь
    available_classes = WeeklySchedule.objects.filter(
        teacher=teacher,
        day_of_week=today_weekday,
        week_start__lte=today,
        week_end__gte=today,
        is_active=True
    ).select_related('course', 'location')
    
    if request.method == 'POST':
        schedule_id = request.POST.get('schedule_id')
        schedule = get_object_or_404(WeeklySchedule, id=schedule_id, teacher=teacher)
        
        # ClassSession үүсгэх
        session = ClassSession.objects.create(
            teacher=teacher,
            course=schedule.course,
            start_time=schedule.start_time,
            form=schedule.form,
            location=schedule.location,
            week_schedule=schedule,
            is_active=True
        )
        
        return redirect('session_qr', token=session.token)
    
    context = {
        'teacher': teacher,
        'available_classes': available_classes,
        'today': today,
    }
    
    return render(request, 'teacher/select_class.html', context)


@login_required
def session_qr(request, token):
    """QR код харуулах хуудас"""
    session = get_object_or_404(ClassSession, token=token)
    
    # Зөвхөн өөрийн хичээлийг харах эрхтэй
    if session.teacher.user != request.user:
        return redirect('teacher_dashboard')
    
    # Ирц бүртгүүлсэн оюутнуудын жагсаалт
    attendances = Attendance.objects.filter(
        session=session
    ).select_related('student').order_by('-timestamp')
    
    # Статистик
    total_enrolled = Enrollment.objects.filter(course=session.course).count()
    attended = attendances.filter(success=True).count()
    failed = attendances.filter(success=False).count()
    
    context = {
        'session': session,
        'attendances': attendances,
        'total_enrolled': total_enrolled,
        'attended': attended,
        'failed': failed,
        'qr_url': request.build_absolute_uri(f'/attendance/scan/{session.token}/'),
    }
    
    return render(request, 'teacher/session_qr.html', context)


@login_required
def close_session(request, token):
    """Хичээл дуусгах, QR код идэвхгүй болгох"""
    session = get_object_or_404(ClassSession, token=token)
    
    if session.teacher.user != request.user:
        return redirect('teacher_dashboard')
    
    session.is_active = False
    session.save()
    
    return redirect('attendance_list', session_id=session.id)


# ============================================
# STUDENT ATTENDANCE VIEWS
# ============================================

def scan_page(request, token):
    """Оюутан QR код уншуулсан хуудас"""
    session = get_object_or_404(ClassSession, token=token)
    
    # Session идэвхтэй эсэхийг шалгах
    if not session.is_active:
        return render(request, 'attendance/session_closed.html', {'session': session})
    
    return render(request, 'attendance/scan_qr.html', {'session': session})


@csrf_exempt
def submit_attendance(request, token):
    """Оюутан ирц бүртгүүлэх API"""
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "POST хүсэлт шаардлагатай."})
    
    session = get_object_or_404(ClassSession, token=token)
    
    # Session идэвхтэй эсэхийг шалгах
    if not session.is_active:
        return JsonResponse({"ok": False, "error": "Хичээл дууссан байна."})
    
    student_code = request.POST.get('student_code')
    lat = request.POST.get('lat')
    lon = request.POST.get('lon')
    
    if not student_code:
        return JsonResponse({"ok": False, "error": "Оюутны код шаардлагатай."})
    
    # Оюутан хайх
    student = Student.objects.filter(student_code=student_code).first()
    if not student:
        return JsonResponse({"ok": False, "error": "Оюутны код олдсонгүй."})
    
    # Хичээлд бүртгэлтэй эсэхийг шалгах
    enrolled = Enrollment.objects.filter(student=student, course=session.course).exists()
    if not enrolled:
        Attendance.objects.create(
            session=session, 
            student=student, 
            lat=lat or None, 
            lon=lon or None, 
            success=False, 
            note="not enrolled"
        )
        return JsonResponse({"ok": False, "error": "Та энэ хичээлд бүртгэлгүй байна."})
    
    # Давхар бүртгэл шалгах
    if Attendance.objects.filter(session=session, student=student).exists():
        return JsonResponse({"ok": False, "error": "Таны ирц аль хэдийн бүртгэгдсэн."})
    
    # GPS байршил шалгах
    try:
        lat_f = float(lat) if lat else None
        lon_f = float(lon) if lon else None
    except (ValueError, TypeError):
        lat_f = None
        lon_f = None
    
    allowed_ok = True
    dist = None
    
    if session.location and (lat_f is not None and lon_f is not None):
        dist = haversine_distance(
            session.location.latitude, 
            session.location.longitude, 
            lat_f, 
            lon_f
        )
        allowed_ok = dist <= session.location.radius_m
    elif session.location:
        allowed_ok = False
    
    # Ирц бүртгэх
    att = Attendance.objects.create(
        session=session, 
        student=student, 
        lat=lat_f, 
        lon=lon_f, 
        success=allowed_ok,
        note="" if allowed_ok else "location not allowed"
    )
    
    if allowed_ok:
        return JsonResponse({
            "ok": True, 
            "message": "Ирц амжилттай бүртгэгдлээ!", 
            "distance_m": int(dist) if dist else None
        })
    else:
        return JsonResponse({
            "ok": False, 
            "error": f"Та зөв байршилд биш байна. ({int(dist) if dist else '?'}м зайтай)", 
            "distance_m": int(dist) if dist else None
        })


# ============================================
# ATTENDANCE LIST & STATISTICS
# ============================================

@login_required
def attendance_list(request, session_id=None):
    """Хичээлийн ирцийн жагсаалт харах"""
    try:
        teacher = request.user.teacherprofile
    except:
        return redirect('teacher_dashboard')
    
    if session_id:
        session = get_object_or_404(ClassSession, id=session_id, teacher=teacher)
        attendances = Attendance.objects.filter(session=session).select_related('student').order_by('-timestamp')
        sessions = [session]
    else:
        # Багшийн бүх хичээлүүд
        sessions = ClassSession.objects.filter(teacher=teacher).order_by('-created_at')
        attendances = None
    
    context = {
        'teacher': teacher,
        'sessions': sessions,
        'attendances': attendances,
        'selected_session': session if session_id else None,
    }
    
    return render(request, 'attendance/list.html', context)


@login_required
def attendance_statistics(request, course_id):
    """Хичээлийн ирцийн статистик"""
    try:
        teacher = request.user.teacherprofile
    except:
        return redirect('teacher_dashboard')
    
    course = get_object_or_404(Course, id=course_id)
    
    # Энэ хичээлийн бүх sessions
    sessions = ClassSession.objects.filter(
        teacher=teacher, 
        course=course
    ).order_by('created_at')
    
    # Бүртгэлтэй оюутнууд
    enrolled_students = Student.objects.filter(
        enrollment__course=course
    ).distinct()
    
    # Оюутан бүрийн ирцийн статистик
    student_stats = []
    for student in enrolled_students:
        attended = Attendance.objects.filter(
            session__in=sessions,
            student=student,
            success=True
        ).count()
        
        failed = Attendance.objects.filter(
            session__in=sessions,
            student=student,
            success=False
        ).count()
        
        total = sessions.count()
        percentage = (attended / total * 100) if total > 0 else 0
        
        student_stats.append({
            'student': student,
            'attended': attended,
            'failed': failed,
            'absent': total - attended - failed,
            'total': total,
            'percentage': round(percentage, 1)
        })
    
    context = {
        'teacher': teacher,
        'course': course,
        'sessions': sessions,
        'student_stats': student_stats,
    }
    
    return render(request, 'attendance/statistics.html', context)


# ============================================
# EXPORT - CSV
# ============================================

@login_required
def export_attendance_csv(request, session_id):
    """Ирцийг CSV файлаар татах"""
    try:
        teacher = request.user.teacherprofile
    except:
        return redirect('teacher_dashboard')
    
    session = get_object_or_404(ClassSession, id=session_id, teacher=teacher)
    attendances = Attendance.objects.filter(session=session).select_related('student')
    
    # CSV үүсгэх
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="attendance_{session.course.code}_{session.created_at.strftime("%Y%m%d")}.csv"'
    
    # UTF-8 BOM нэмэх (Excel-д монгол үсэг зөв харуулахын тулд)
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Header
    writer.writerow([
        'Оюутны код',
        'Оюутны нэр',
        'Ирсэн цаг',
        'Байршил (Latitude)',
        'Байршил (Longitude)',
        'Зай (м)',
        'Төлөв',
        'Тэмдэглэл'
    ])
    
    # Data
    for att in attendances:
        distance = ''
        if att.lat and att.lon and session.location:
            distance = str(int(haversine_distance(
                session.location.latitude, 
                session.location.longitude, 
                att.lat, 
                att.lon
            )))
        
        writer.writerow([
            att.student.student_code,
            att.student.full_name,
            att.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            att.lat or '',
            att.lon or '',
            distance,
            'Ирсэн' if att.success else 'Тасалсан',
            att.note
        ])
    
    # Тайлан хадгалах
    AttendanceReport.objects.create(
        teacher=teacher,
        course=session.course,
        session=session,
        format='csv'
    )
    
    return response


# ============================================
# EXPORT - EXCEL
# ============================================

@login_required
def export_attendance_excel(request, session_id):
    """Ирцийг Excel файлаар татах"""
    try:
        teacher = request.user.teacherprofile
    except:
        return redirect('teacher_dashboard')
    
    session = get_object_or_404(ClassSession, id=session_id, teacher=teacher)
    attendances = Attendance.objects.filter(session=session).select_related('student')
    
    # Excel workbook үүсгэх
    wb = Workbook()
    ws = wb.active
    ws.title = "Ирцийн мэдээлэл"
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Гарчиг мэдээлэл
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Ирцийн мэдээлэл - {session.course.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Багш: {session.teacher.name} | Огноо: {session.created_at.strftime('%Y-%m-%d')} | Цаг: {session.start_time}"
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Header row
    headers = [
        'Оюутны код',
        'Оюутны нэр',
        'Ирсэн цаг',
        'Latitude',
        'Longitude',
        'Зай (м)',
        'Төлөв',
        'Тэмдэглэл'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for row_idx, att in enumerate(attendances, start=5):
        distance = ''
        if att.lat and att.lon and session.location:
            distance = int(haversine_distance(
                session.location.latitude, 
                session.location.longitude, 
                att.lat, 
                att.lon
            ))
        
        data = [
            att.student.student_code,
            att.student.full_name,
            att.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            att.lat or '',
            att.lon or '',
            distance,
            'Ирсэн' if att.success else 'Тасалсан',
            att.note
        ]
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if col_idx in [2, 8] else "center")
            
            # Төлөв багананд өнгө өгөх
            if col_idx == 7:
                if att.success:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006")
    
    # Баганын өргөнийг тохируулах
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20
    
    # Статистик нэмэх
    total = attendances.count()
    attended = attendances.filter(success=True).count()
    failed = attendances.filter(success=False).count()
    
    stats_row = len(attendances) + 6
    ws.cell(row=stats_row, column=1, value="Нийт:").font = Font(bold=True)
    ws.cell(row=stats_row, column=2, value=total)
    ws.cell(row=stats_row+1, column=1, value="Ирсэн:").font = Font(bold=True)
    ws.cell(row=stats_row+1, column=2, value=attended)
    ws.cell(row=stats_row+2, column=1, value="Тасалсан:").font = Font(bold=True)
    ws.cell(row=stats_row+2, column=2, value=failed)
    
    # Response үүсгэх
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="attendance_{session.course.code}_{session.created_at.strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    
    # Тайлан хадгалах
    AttendanceReport.objects.create(
        teacher=teacher,
        course=session.course,
        session=session,
        format='excel'
    )
    
    return response


# ============================================
# EXPORT - PDF
# ============================================

@login_required
def export_attendance_pdf(request, session_id):
    """Ирцийг PDF файлаар татах"""
    try:
        teacher = request.user.teacherprofile
    except:
        return redirect('teacher_dashboard')
    
    session = get_object_or_404(ClassSession, id=session_id, teacher=teacher)
    attendances = Attendance.objects.filter(session=session).select_related('student')
    
    # PDF buffer үүсгэх
    buffer = io.BytesIO()
    
    # PDF document үүсгэх
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    # Story container
    elements = []
    styles = getSampleStyleSheet()
    
    # Гарчиг
    title_style = styles['Heading1']
    title_style.alignment = 1  # Center
    title = Paragraph(f"Ирцийн мэдээлэл - {session.course.name}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Мэдээлэл
    info_style = styles['Normal']
    info_style.alignment = 1
    info_text = f"Багш: {session.teacher.name} | Огноо: {session.created_at.strftime('%Y-%m-%d')} | Цаг: {session.start_time}"
    info = Paragraph(info_text, info_style)
    elements.append(info)
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['№', 'Оюутны код', 'Оюутны нэр', 'Ирсэн цаг', 'Зай (м)', 'Төлөв']]
    
    for idx, att in enumerate(attendances, start=1):
        distance = ''
        if att.lat and att.lon and session.location:
            distance = str(int(haversine_distance(
                session.location.latitude, 
                session.location.longitude, 
                att.lat, 
                att.lon
            )))
        
        data.append([
            str(idx),
            att.student.student_code,
            att.student.full_name,
            att.timestamp.strftime('%Y-%m-%d %H:%M'),
            distance,
            'Ирсэн' if att.success else 'Тасалсан'
        ])
    
    # Table үүсгэх
    table = Table(data, colWidths=[0.5*inch, 1.2*inch, 2.5*inch, 1.5*inch, 1*inch, 1*inch])
    
    # Table style
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Статистик
    total = attendances.count()
    attended = attendances.filter(success=True).count()
    failed = attendances.filter(success=False).count()
    
    stats_text = f"Нийт: {total} | Ирсэн: {attended} | Тасалсан: {failed}"
    stats = Paragraph(stats_text, styles['Normal'])
    elements.append(stats)
    
    # PDF үүсгэх
    doc.build(elements)
    
    # Response
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="attendance_{session.course.code}_{session.created_at.strftime("%Y%m%d")}.pdf"'
    response.write(pdf)
    
    # Тайлан хадгалах
    AttendanceReport.objects.create(
        teacher=teacher,
        course=session.course,
        session=session,
        format='pdf'
    )
    
    return response


# ============================================
# EXPORT - COURSE ATTENDANCE COMBINED
# ============================================

@login_required
def export_course_attendance_excel(request, course_id):
    """Хичээлийн бүх хичээлүүдийн ирцийг нэгтгэсэн Excel"""
    try:
        teacher = request.user.teacherprofile
    except:
        return redirect('teacher_dashboard')
    
    course = get_object_or_404(Course, id=course_id)
    sessions = ClassSession.objects.filter(
        teacher=teacher, 
        course=course
    ).order_by('created_at')
    
    students = Student.objects.filter(
        enrollment__course=course
    ).distinct().order_by('student_code')
    
    # Excel workbook үүсгэх
    wb = Workbook()
    ws = wb.active
    ws.title = "Ирцийн нэгтгэл"
    
    # Гарчиг
    ws.merge_cells('A1:Z1')
    ws['A1'] = f"Ирцийн нэгтгэл - {course.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Headers
    ws.cell(row=3, column=1, value="Оюутны код")
    ws.cell(row=3, column=2, value="Оюутны нэр")
    
    for col_idx, session in enumerate(sessions, start=3):
        ws.cell(row=3, column=col_idx, value=session.created_at.strftime('%m/%d'))
    
    ws.cell(row=3, column=len(sessions)+3, value="Нийт")
    ws.cell(row=3, column=len(sessions)+4, value="Ирсэн")
    ws.cell(row=3, column=len(sessions)+5, value="Хувь")
    
    # Data
    for row_idx, student in enumerate(students, start=4):
        ws.cell(row=row_idx, column=1, value=student.student_code)
        ws.cell(row=row_idx, column=2, value=student.full_name)
        
        attended_count = 0
        for col_idx, session in enumerate(sessions, start=3):
            att = Attendance.objects.filter(session=session, student=student).first()
            if att:
                if att.success:
                    ws.cell(row=row_idx, column=col_idx, value="✓")
                    attended_count += 1
                else:
                    ws.cell(row=row_idx, column=col_idx, value="✗")
            else:
                ws.cell(row=row_idx, column=col_idx, value="-")
        
        total = len(sessions)
        percentage = (attended_count / total * 100) if total > 0 else 0
        
        ws.cell(row=row_idx, column=len(sessions)+3, value=total)
        ws.cell(row=row_idx, column=len(sessions)+4, value=attended_count)
        ws.cell(row=row_idx, column=len(sessions)+5, value=f"{percentage:.1f}%")
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="course_attendance_{course.code}.xlsx"'
    wb.save(response)
    
    return response
